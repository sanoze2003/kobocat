# coding: utf-8
import os
import time
import csv
import tempfile
from io import BytesIO

from django.urls import reverse
from django.core.files.storage import get_storage_class, FileSystemStorage
from django.utils import timezone
from openpyxl import load_workbook

from onadata.apps.viewer.models.export import Export
from onadata.apps.viewer.views import kml_export, export_download
from onadata.libs.utils.export_tools import generate_export
from onadata.libs.utils.user_auth import http_auth_string
from .test_base import TestBase


class TestFormExports(TestBase):

    def setUp(self):
        TestBase.setUp(self)
        self._create_user_and_login()
        self._publish_transportation_form_and_submit_instance()
        self.csv_url = reverse('csv_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string})
        self.xls_url = reverse('xls_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string})

    def _num_rows(self, content, export_format):
        def xls_rows(f):
            wb = load_workbook(BytesIO(f))
            return wb[wb.sheetnames[0]].max_row

        def csv_rows(f):
            return len([line for line in csv.reader(f.decode().strip().split('\n'))])
        num_rows_fn = {
            'xls': xls_rows,
            'csv': csv_rows,
        }
        return num_rows_fn[export_format](content)

    def test_csv_raw_export_name(self):
        response = self.client.get(self.csv_url + '?raw=1')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Disposition'], 'attachment;')

    def _filter_export_test(self, url, export_format):
        """
        Test filter exports.  Use sleep to ensure we don't have unique seconds.
        Number of rows equals number of surveys plus 1, the header row.
        """
        time.sleep(1)
        # 1 survey exists before this time
        start_time = timezone.now().strftime('%y_%m_%d_%H_%M_%S')
        time.sleep(1)
        s = self.surveys[1]
        self._make_submission(
            os.path.join(self.this_directory, 'fixtures',
                         'transportation', 'instances', s, s + '.xml'))
        time.sleep(1)
        # 2 surveys exist before this time
        end_time = timezone.now().strftime('%y_%m_%d_%H_%M_%S')
        time.sleep(1)
        # 3 surveys exist in total
        s = self.surveys[2]
        self._make_submission(
            os.path.join(self.this_directory, 'fixtures',
                         'transportation', 'instances', s, s + '.xml'))
        # test restricting to before end time
        params = {'end': end_time}
        response = self.client.get(url, params)
        self.assertEqual(response.status_code, 200)
        content = self._get_response_content(response)
        self.assertEqual(self._num_rows(content, export_format), 3)
        # test restricting to after start time, thus excluding the initial
        # submission
        params = {'start': start_time}
        response = self.client.get(url, params)
        self.assertEqual(response.status_code, 200)
        content = self._get_response_content(response)
        self.assertEqual(self._num_rows(content, export_format), 3)
        # test no time restriction
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        content = self._get_response_content(response)
        self.assertEqual(self._num_rows(content, export_format), 4)
        # test restricting to between start time and end time
        params = {'start': start_time, 'end': end_time}
        response = self.client.get(url, params)
        self.assertEqual(response.status_code, 200)
        content = self._get_response_content(response)
        self.assertEqual(self._num_rows(content, export_format), 2)

    def test_filter_by_date_csv(self):
        self._filter_export_test(self.csv_url, 'csv')

    def test_filter_by_date_xls(self):
        self._filter_export_test(self.xls_url, 'xls')

    def test_restrict_csv_export_if_not_shared(self):
        response = self.anon.get(self.csv_url)
        self.assertEqual(response.status_code, 403)

    def test_xls_raw_export_name(self):
        response = self.client.get(self.xls_url + '?raw=1')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Disposition'], 'attachment;')

    def test_restrict_xls_export_if_not_shared(self):
        response = self.anon.get(self.xls_url)
        self.assertEqual(response.status_code, 403)

    def test_restrict_kml_export_if_not_shared(self):
        url = reverse(kml_export, kwargs={'username': self.user.username,
                                          'id_string': self.xform.id_string})
        response = self.anon.get(url)
        self.assertEqual(response.status_code, 403)

    def test_allow_csv_export_if_shared(self):
        self.xform.shared_data = True
        self.xform.save()
        response = self.anon.get(self.csv_url)
        self.assertEqual(response.status_code, 200)

    def test_allow_xls_export_if_shared(self):
        self.xform.shared_data = True
        self.xform.save()
        response = self.anon.get(self.xls_url)
        self.assertEqual(response.status_code, 200)

    def test_allow_kml_export_if_shared(self):
        self.xform.shared_data = True
        self.xform.save()
        url = reverse(kml_export, kwargs={'username': self.user.username,
                                          'id_string': self.xform.id_string})
        response = self.anon.get(url)
        self.assertEqual(response.status_code, 200)

    def test_allow_csv_export(self):
        response = self.client.get(self.csv_url)
        self.assertEqual(response.status_code, 200)

    def test_allow_xls_export(self):
        response = self.client.get(self.xls_url)
        self.assertEqual(response.status_code, 200)

    def test_allow_kml_export(self):
        url = reverse(kml_export, kwargs={'username': self.user.username,
                                          'id_string': self.xform.id_string})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_allow_csv_export_for_basic_auth(self):
        extra = {
            'HTTP_AUTHORIZATION': http_auth_string(self.login_username,
                                                   self.login_password)
        }
        response = self.anon.get(self.csv_url, **extra)
        self.assertEqual(response.status_code, 200)

    def test_allow_xls_export_for_basic_auth(self):
        extra = {
            'HTTP_AUTHORIZATION': http_auth_string(self.login_username,
                                                   self.login_password)
        }
        response = self.anon.get(self.xls_url, **extra)
        self.assertEqual(response.status_code, 200)

    def test_allow_kml_export_for_basic_auth(self):
        extra = {
            'HTTP_AUTHORIZATION': http_auth_string(self.login_username,
                                                   self.login_password)
        }
        url = reverse(kml_export, kwargs={'username': self.user.username,
                                          'id_string': self.xform.id_string})
        response = self.anon.get(url, **extra)
        self.assertEqual(response.status_code, 200)

    def test_allow_export_download_for_basic_auth(self):
        extra = {
            'HTTP_AUTHORIZATION': http_auth_string(self.login_username,
                                                   self.login_password)
        }
        # create export
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        self.assertTrue(isinstance(export, Export))
        url = reverse(export_download, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': export.export_type,
            'filename': export.filename
        })
        response = self.anon.get(url, **extra)
        default_storage = get_storage_class()()
        if not isinstance(default_storage, FileSystemStorage):
            self.assertEqual(response.status_code, 302)
        else:
            self.assertEqual(response.status_code, 200)
